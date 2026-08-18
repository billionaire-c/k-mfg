"""
한국산업단지공단 공개 통계 → public/data/industrial-parks.json

입력(기본: Downloads):
  - 전국산업단지현황통계.xlsx  (단지 마스터)
  - 국가산업단지 산업동향정보.xlsx (가동·생산·고용·가동률)

좌표는 단지 상세 주소가 없어 시도 중심점으로 매핑합니다.
"""

from __future__ import annotations

import json
import math
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "public" / "data" / "industrial-parks.json"

DEFAULT_STATUS = Path.home() / "Downloads" / "한국산업단지공단_전국산업단지현황통계_20251231.xlsx"
DEFAULT_TREND = Path.home() / "Downloads" / "한국산업단지공단_국가산업단지 산업동향정보_20260331.xlsx"

# 시도 대표 좌표 (지도 포커스용 근사치)
SIDO_CENTERS: dict[str, tuple[float, float]] = {
    "서울": (37.5665, 126.9780),
    "부산": (35.1796, 129.0756),
    "대구": (35.8714, 128.6014),
    "인천": (37.4563, 126.7052),
    "광주": (35.1595, 126.8526),
    "대전": (36.3504, 127.3845),
    "울산": (35.5384, 129.3114),
    "세종": (36.4801, 127.2890),
    "경기": (37.4138, 127.5183),
    "강원": (37.8228, 128.1555),
    "충북": (36.6357, 127.4914),
    "충남": (36.5184, 126.8000),
    "전북": (35.7175, 127.1530),
    "전남": (34.8679, 126.9910),
    "경북": (36.4919, 128.8889),
    "경남": (35.4606, 128.2132),
    "제주": (33.4996, 126.5312),
}

# 동향 약칭 → (선호 시도, 단지명에 포함되면 매칭되는 키워드들)
TREND_MATCH: dict[str, tuple[str | None, list[str]]] = {
    "서울": ("서울", ["한국수출산업"]),
    "녹산": ("부산", ["녹산"]),
    "대구": ("대구", ["대구국가", "대구"]),
    "남동": ("인천", ["남동"]),
    "부평": ("인천", ["부평"]),  # 단지명이 한국수출산업일 수 있어 시군으로 보정
    "주안": ("인천", ["주안"]),
    "광주첨단": ("광주", ["광주첨단"]),
    "빛그린": ("광주", ["빛그린"]),
    "온산": ("울산", ["온산"]),
    "울산ㆍ미포": ("울산", ["울산", "미포"]),
    "반월": ("경기", ["반월"]),
    "용인첨단시스템반도체": ("경기", ["용인첨단"]),
    "시화": ("경기", ["시화"]),
    "시화MTV": ("경기", ["시화MTV", "시화 MTV", "MTV"]),
    "송산그린시티": ("경기", ["송산"]),
    "파주탄현": ("경기", ["파주탄현", "탄현"]),
    "동두천": ("경기", ["동두천"]),
    "북평": ("강원", ["북평"]),
    "오송생명과학": ("충북", ["오송생명"]),
    "석문": ("충남", ["석문"]),
    "아산": ("충남", ["아산국가", "아산"]),
    "장항생태": ("충남", ["장항"]),
    "국가식품클러스터": ("전북", ["국가식품클러스터"]),
    "국가식품클러스터(외)": ("전북", ["국가식품클러스터"]),
    "군산": ("전북", ["군산"]),
    "군산2": ("전북", ["군산2", "군장"]),
    "익산": ("전북", ["익산"]),
    "광양": ("전남", ["광양"]),
    "대불": ("전남", ["대불"]),
    "대불(외)": ("전남", ["대불"]),
    "여수": ("전남", ["여수"]),
    "구미": ("경북", ["구미국가", "구미"]),
    "구미(외)": ("경북", ["구미하이테크", "구미"]),
    "포항": ("경북", ["포항국가", "포항"]),
    "포항블루밸리": ("경북", ["블루밸리"]),
    "경남항공": ("경남", ["경남항공", "항공"]),
    "밀양나노": ("경남", ["밀양나노", "나노"]),
    "안정": ("경남", ["안정"]),
    "진해": ("경남", ["진해"]),
    "창원": ("경남", ["창원"]),
}


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v) if not isinstance(v, int) else v
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return None


def _clean_name(name: str) -> str:
    return re.sub(r"\s+", " ", name).strip()


def load_parks(path: Path) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["전국산업단지현황(상위단지)"]
    parks: list[dict] = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        park_type = row[0]
        sido = row[1]
        sigungu = row[2]
        name = row[3]
        if not park_type or not sido or not name:
            continue
        if str(park_type).strip() in ("합계", "계", "총계"):
            continue
        park_type = str(park_type).strip()
        sido = str(sido).strip()
        sigungu = str(sigungu).strip() if sigungu else ""
        name = _clean_name(str(name))
        status = str(row[4]).strip() if row[4] else ""
        designated = _num(row[5])
        managed = _num(row[6])
        facility_total = _num(row[7])
        sale_target = _num(row[8])
        sold = _num(row[9])
        unsold = _num(row[10])
        sale_rate = _num(row[11])
        tenants = _num(row[12])
        operating = _num(row[13])
        latlng = SIDO_CENTERS.get(sido)
        if not latlng:
            continue
        parks.append(
            {
                "id": f"{sido}-{sigungu}-{name}",
                "type": park_type,
                "sido": sido,
                "sigungu": sigungu,
                "name": name,
                "status": status,
                "designatedArea": designated,
                "managedArea": managed,
                "facilityArea": facility_total,
                "saleTargetArea": sale_target,
                "soldArea": sold,
                "unsoldArea": unsold,
                "saleRate": sale_rate,
                "tenants": int(tenants) if tenants is not None else None,
                "operating": int(operating) if operating is not None else None,
                "lat": latlng[0],
                "lng": latlng[1],
                "address": f"{sido} {sigungu}".strip(),
                "trendKey": None,
            }
        )

    # 요약 시트: (1) 조성·분양 / (2) 입주·가동
    summary: dict = {"parkCount": 1359, "byType": {}}
    try:
        sw = wb["요약"]
        # (1) 조성 및 분양 — 단지유형별 단지수·분양률
        for row in sw.iter_rows(min_row=6, max_row=10, values_only=True):
            label = str(row[0]).strip() if row[0] else ""
            if label not in ("국가", "일반", "도시첨단", "농공", "총합"):
                continue
            summary["byType"][label] = {
                "parks": int(_num(row[1]) or 0),
                "designatedArea": _num(row[2]),
                "saleRate": _num(row[8]),
                "tenants": None,
                "operating": None,
            }
            if label == "총합":
                summary["parkCount"] = int(_num(row[1]) or 0)
        # (2) 입주 및 고용
        for row in sw.iter_rows(min_row=17, max_row=21, values_only=True):
            label = str(row[0]).strip() if row[0] else ""
            if label not in summary["byType"]:
                continue
            summary["byType"][label]["tenants"] = int(_num(row[1]) or 0)
            summary["byType"][label]["operating"] = int(_num(row[2]) or 0)
        # 총합 입주 = 유형 합
        if "총합" in summary["byType"]:
            tenants_sum = 0
            operating_sum = 0
            for key in ("국가", "일반", "도시첨단", "농공"):
                block = summary["byType"].get(key) or {}
                if block.get("tenants"):
                    tenants_sum += block["tenants"]
                if block.get("operating"):
                    operating_sum += block["operating"]
            summary["byType"]["총합"]["tenants"] = tenants_sum or None
            summary["byType"]["총합"]["operating"] = operating_sum or None
    except Exception:
        # 파싱 실패 시 리스트에서 집계
        from collections import Counter

        c = Counter(p["type"] for p in parks)
        summary["byType"] = {
            k: {"parks": v, "tenants": None, "operating": None, "saleRate": None, "designatedArea": None}
            for k, v in c.items()
        }
        summary["byType"]["총합"] = {
            "parks": len(parks),
            "tenants": None,
            "operating": None,
            "saleRate": None,
            "designatedArea": None,
        }

    return parks, summary


def load_trends(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet_rows(name: str, start: int):
        ws = wb[name]
        out = {}
        for row in ws.iter_rows(min_row=start, values_only=True):
            key = row[0]
            if not key:
                continue
            key = str(key).strip()
            if key in ("계", "합계", "총계") or key.startswith("계"):
                break
            out[key] = row
        return out

    tenants = sheet_rows("표1 단지별 입주", 5)
    prod = sheet_rows("표4 단지별 생산", 4)
    emp = sheet_rows("표8 단지별 고용", 5)
    util = sheet_rows("표10 단지별 가동률", 4)

    keys = list(tenants.keys())
    trends = []
    for key in keys:
        t = tenants.get(key)
        p = prod.get(key)
        e = emp.get(key)
        u = util.get(key)
        sido_hint, _ = TREND_MATCH.get(key, (None, []))
        latlng = SIDO_CENTERS.get(sido_hint) if sido_hint else None
        trends.append(
            {
                "key": key,
                "sido": sido_hint,
                "tenants": int(_num(t[2]) or 0) if t else None,
                "operating": int(_num(t[4]) or 0) if t else None,
                "production": round(float(_num(p[2])), 1) if p and _num(p[2]) is not None else None,
                "productionQoQ": round(float(_num(p[5])), 2) if p and _num(p[5]) is not None else None,
                "employment": int(_num(e[2]) or 0) if e else None,
                "employmentQoQ": round(float(_num(e[6])), 2) if e and _num(e[6]) is not None else None,
                "utilization": round(float(_num(u[5])), 2) if u and _num(u[5]) is not None else None,
                "utilizationQoQ": round(float(_num(u[7])), 2) if u and _num(u[7]) is not None else None,
                "lat": latlng[0] if latlng else None,
                "lng": latlng[1] if latlng else None,
            }
        )
    return trends


def attach_trend_keys(parks: list[dict], trends: list[dict]) -> None:
    used: set[str] = set()
    for trend in trends:
        key = trend["key"]
        sido_hint, keywords = TREND_MATCH.get(key, (None, [key]))
        candidates = []
        for park in parks:
            if park["type"] != "국가":
                continue
            if sido_hint and park["sido"] != sido_hint:
                # 빛그린·아산·명지녹산 등 2시도 걸침: 힌트 시도만 우선
                if key not in ("빛그린", "아산", "녹산"):
                    continue
            hay = f"{park['name']} {park['sigungu']}"
            score = 0
            for kw in keywords:
                if kw and kw in hay:
                    score += len(kw)
            if key == "부평" and park["sigungu"] == "부평구":
                score += 20
            if key == "서울" and park["sido"] == "서울" and "수출" in park["name"]:
                score += 20
            if key == "주안" and ("미추홀" in park["sigungu"] or "남구" in park["sigungu"]):
                score += 5
            if score > 0:
                candidates.append((score, park))
        if not candidates:
            continue
        candidates.sort(key=lambda x: -x[0])
        best = candidates[0][1]
        # 동일 키 중복 방지: 최고점만
        if key in used:
            continue
        best["trendKey"] = key
        used.add(key)


def build_regions(parks: list[dict]) -> list[dict]:
    buckets: dict[str, dict] = {}
    for park in parks:
        sido = park["sido"]
        if sido not in buckets:
            lat, lng = SIDO_CENTERS[sido]
            buckets[sido] = {
                "id": f"sido-{sido}",
                "sido": sido,
                "name": f"{sido} 산업단지",
                "address": sido,
                "lat": lat,
                "lng": lng,
                "parkCount": 0,
                "tenants": 0,
                "operating": 0,
                "byType": {},
            }
        b = buckets[sido]
        b["parkCount"] += 1
        if park["tenants"]:
            b["tenants"] += park["tenants"]
        if park["operating"]:
            b["operating"] += park["operating"]
        b["byType"][park["type"]] = b["byType"].get(park["type"], 0) + 1
    return sorted(buckets.values(), key=lambda x: -x["parkCount"])


def main() -> None:
    status_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_STATUS
    trend_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_TREND
    if not status_path.exists():
        raise SystemExit(f"현황 파일 없음: {status_path}")
    if not trend_path.exists():
        raise SystemExit(f"동향 파일 없음: {trend_path}")

    parks, summary = load_parks(status_path)
    trends = load_trends(trend_path)
    attach_trend_keys(parks, trends)
    regions = build_regions(parks)

    matched = sum(1 for p in parks if p.get("trendKey"))
    payload = {
        "source": "한국산업단지공단_전국산업단지현황통계 · 국가산업단지 산업동향정보",
        "sourceUrl": "https://www.data.go.kr/tcs/dss/selectDataSetList.do?keyword=%EC%82%B0%EC%97%85%EB%8B%A8%EC%A7%80&conditionType=search",
        "statusAsOf": "2025-12-31",
        "trendAsOf": "2026-03-31",
        "updatedAt": "2025-12-31",
        "note": "단지별 상세 주소가 공개되지 않아 지도 좌표는 시도 중심점으로 표시합니다. 국가산단 동향은 공단 관할 주요 단지 기준입니다.",
        "summary": summary,
        "totalParks": len(parks),
        "trendMatched": matched,
        "regions": regions,
        "parks": parks,
        "trends": trends,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {OUT} parks={len(parks)} regions={len(regions)} trends={len(trends)} matched={matched}")


if __name__ == "__main__":
    main()
